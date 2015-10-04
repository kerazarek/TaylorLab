#!/usr/bin/env python

from __future__ import print_function
from openpyxl import load_workbook
import csv, re, sys
from numpy import mean


script, dock = sys.argv

###
# 	v1.0 10/2/15 10:57pm


class Xlsx:
	def __init__(self, address):
		self.address = address
		self.workbook = load_workbook(self.address)
		self.sheets_list = self.workbook.get_sheet_names()

	def sheet(self, sheet):
		self.active_sheet = self.workbook.get_sheet_by_name(sheet)
# 		print(self.active_sheet.columns)

		self.col_keys = {}
		self.col_heads = {}
		for col in self.active_sheet.columns:
			for c in col:
				 if c.row == 1:
					self.col_keys[str(c.value)] = c.column
					self.col_heads[c.column] = str(c.value)

		self.row_keys = {}
		self.row_heads = {}
		for col in self.active_sheet.columns:
			for c in col:
				 if c.column == 'A' and c.row != 1 and c.value != None:
					self.row_keys[str(c.value)] = c.row
					self.row_heads[c.row] = str(c.value)


	def look_up(self, querysheet, queryrow, querycol):
		self.sheet(querysheet)

		rkey = self.row_keys[queryrow]
		ckey = self.col_keys[querycol]
		caddress = ckey+str(rkey)
		cell = self.active_sheet.cell(caddress)
		contents = cell.value
		return contents


	def sheet_dic(self, sheet):
		self.sheet(sheet)
		self.sheet_dic = {}
		for row in self.active_sheet.rows:
			if row[0].value != None and row[0].row != 1:
				row_head = str(row[0].value)
				row_data = {}
				for c in row:
					if c.column != 'A' and c.row != 1 and c.value != None:
						col_head = self.col_heads[c.column]
						if type(c.value) is (int or float):
							row_data[col_head] = int(c.value)
						else:
							row_data[col_head] = str(c.value)
				self.sheet_dic[row_head] = row_data
		return self.sheet_dic

class Csv:
	def __init__(self, source_csv):
		self.source_csv = source_csv
		self.dic = {}

		csv_open = open(self.source_csv, 'r')
		csv_read = csv.reader(csv_open, delimiter=',', quotechar='"')

		csv_list = []
		for row in csv_read:
			csv_list.append(row)

		keys = csv_list[0]
		for row in csv_list:
			dic_entry = {}
			for i in range(0, len(keys)):
				if row[i] != '':
					dic_entry[keys[i]] = row[i]
			self.dic[row[0]] = dic_entry

# class Pvrd():
# 	def __init__(self, address):
# 		self.pdbqt_str = open(address).read()
# 		str = self.pdbqt_str
# 		for line in open(address).read():
# 			if re.search( 'HETATM', str ):
# 				print(line)

class Docking:
	def basic_params(self):
		if dock[:1] == "h":
			self.prot = "hepi"
		elif dock[:1] == "p":
			self.prot = "p300"
		else:
			print("!!! BAD DOCK ID !!!")

		self.dock = dock
		self.date = str(docks_xlsx.look_up(self.prot, dock, "DATE"))
		self.prot = str(docks_xlsx.look_up(self.prot, dock, "PROT"))
		self.specprot = str(docks_xlsx.look_up(self.prot, dock, "SPECPROT"))
		self.ligset = str(docks_xlsx.look_up(self.prot, dock, "LIGSET"))
		self.box = str(docks_xlsx.look_up(self.prot, dock, "BOX"))
		self.exhaust = str(docks_xlsx.look_up(self.prot, dock, "EXHAUST"))
		self.n_models = str(docks_xlsx.look_up(self.prot, dock, "n_MODELS"))
		self.n_cpus = str(docks_xlsx.look_up(self.prot, dock, "n_CPUS"))
		self.notes = str(docks_xlsx.look_up(self.prot, dock, "notes"))

		self.params = {}
		for bp in basic_params:
			evalbp = "self."+bp
			self.params[bp] = eval(evalbp)

	def ligset_params(self):
		self.ligset_list_sh = str(docks_xlsx.look_up("ligsets", self.ligset, "lig_list"))
		self.ligset_list = sh_list_to_py(self.ligset_list_sh)

	def box_params(self):
		self.box_params = {}
		self.box_params['name'] = str(docks_xlsx.look_up('gridboxes', self.box, 'name'))
		self.box_params['description'] = str(docks_xlsx.look_up('gridboxes', self.box, 'description'))
		self.box_params['size_x'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_x'))
		self.box_params['size_y'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_y'))
		self.box_params['size_z'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_z'))
		self.box_params['center_x'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_x'))
		self.box_params['center_y'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_y'))
		self.box_params['center_z'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_z'))
		self.box_params['notes'] = str(docks_xlsx.look_up('gridboxes', self.box, 'notes'))
		self.box_params['size_tuple'] = self.box_params['size_x'], self.box_params['size_y'], self.box_params['size_z']
		self.box_params['center_tuple'] = self.box_params['center_x'], self.box_params['center_y'], self.box_params['center_z']
		self.box_params['name'] = str(docks_xlsx.look_up('gridboxes', self.box, 'name'))

	def binding_site_params(self):
		self.baseprot = docks_xlsx.look_up("pdbs", self.specprot, "baseprot")
		self.bs_list_sh = docks_xlsx.look_up("pdbs", self.specprot, "binding_sites")
		self.bs_list = sh_list_to_py(self.bs_list_sh)
		if self.prot == "p300":
			self.bs_list = ["lys", "side", "coa_ado", "coa_adpp", "coa_pant", "allo1", "allo2"]
# 		self.bs_list = ["ado", "pant", "side", "lys", "allo1", "allo2"] # !!placholder
# 		self.bs_list = ["adph", "fdla", "allo"] # !!placholder
		self.lig_pdbs_dic = {}
		self.lig_csvs_dic = {}
		self.lig_csvdics_dic = {}
		self.a5res_pdbs_dic = {}
		self.a5res_csvs_dic = {}
		self.a5res_csvdics_dic = {}
		self.a5res_lists_dic = {}
		self.a5res_atom_lists_dic = {}
# 		bs_dic_addresses = {
# 			self.lig_pdbs_dic: self.bindingsites_dir+"lig_pdbs/"+self.baseprot+"_bs_"+bs
# 			self.lig_csvs_dic: self.bindingsites_dir+"lig_csvs/"+self.baseprot+"_bs_"+bs
# # 			self.lig_csvdics_dic: self.bindingsites_dir+"lig_csvdics/"+self.baseprot+"_bs_"+bs
# 			self.a5res_pdbs_dic: self.bindingsites_dir+"a5res_pdbs/"+self.baseprot+"_bs_"+bs
# 			self.a5res_csvs_dic: self.bindingsites_dir+"a5res_csvs/"+self.baseprot+"_bs_"+bs
# # 			self.a5res_csvdics_dic: self.bindingsites_dir+"a5res_csvdics/"+self.baseprot+"_bs_"+bs
# 			self.a5res_lists_dic: self.bindingsites_dir+"a5res_lists/"+self.baseprot+"_bs_"+bs
# 			self.a5res_atom_lists_dic: self.bindingsites_dir+"a5res_atom_lists/"+self.baseprot+"_bs_"+bs
# 		}

	def file_addresses(self): # ALL DIRS END IN /
		# Basic Directories
		self.lab_dir = "/Users/zarek/lab/"
		self.docking_dir = self.lab_dir+"Docking/"
		self.prot_dir = self.docking_dir+self.prot+"/"
		self.bindingsites_dir = self.docking_dir+"binding_sites/"+self.baseprot+"/"
		self.ligset_dir = self.docking_dir+"ligsets/"+self.ligset+"/"
		self.dock_dir = self.prot_dir+self.dock+"/"
		self.res_dir = self.dock_dir+"results/"
		self.pvrd_pdbqts_dir = self.dock_dir+"pvrd_pdbqts/"
		# Input Files
# 		self.res_pdbqt = "{}results/{}_{}_results.pdbqt".format(self.dock_dir, self.dock, self
# 		self.res_pdbqt = "{}results/{}_{}_m{}.pdbqt"
		# Output Files
		self.alldata_csv_address = self.prot_dir+self.dock+"/"+self.dock+"_alldata.csv"
		self.post_analysis_csv_address = self.prot_dir+self.dock+"/"+self.dock+"_analysis.csv"

	def fetch_params(self):
		self.basic_params()
		self.ligset_params()
		self.box_params()
		self.binding_site_params()
		self.file_addresses()

	def __init__(self, dock):
		self.dock = dock
		self.fetch_params()

	def list_param(self, param):
		exec_print_param = "print(self."+param+")"
		exec(exec_print_param)

	def load_vina_results(self):
		for lig in self.ligset_list:
			print(lig)
			res_pdbqt = self.res_dir+self.dock+"_"+lig+"_results.pdbqt"
			for m in range(1, (int(self.n_models) + 1)):
				try:
					pvrd_pdbqt = str(self.pvrd_pdbqts_dir)+self.dock+"_"+lig+"_m"+str(m)+".pdbqt"
					pvrd = Pvrd(pvrd_pdbqt)
# 					print(pvrd.dic)
				except IOError:
					pass

	def generate_alldata_dic(self):
		self.alldata_dic = Csv(self.alldata_csv_address).dic

	def alldata_lig_subsets_dic(self):
		self.generate_alldata_dic()
		dic = {}
		for lig in self.ligset_list:
			dic[lig] = dic_sort(self.alldata_dic, "LIG", match = lig)
		return dic

# 	is this one even useful?
	def alldata_lig_E_lists_dic(self):
		dic = {}
		for lig, subset_dic in self.alldata_lig_subsets_dic().items():
			lig_E_list = []
			for k, v in subset_dic.items():
				lig_E_list.append(float(v['E']))
			dic[lig] = lig_E_list
		return dic

# 	def bs_lig_E_list_dic(self):
# 		dic = {}
# 		for bs in self.bs_list:
# 			dic[bs] = {}
# # 			score_header = "resis_score_fraction_"+bs
# 			for lig, subset_dic in self.alldata_lig_subsets_dic().items():
# 				lig_E_list = []
# 				for k, v in subset_dic.items():
# 					lig_E_list.append(v['E'])
# 				dic[bs][lig] = lig_E_list
# 		return dic

	def alldata_assign_bs(self, threshold = 0.10):
		self.generate_alldata_dic()
		new_alldata_dic = {}
		for k, data in self.alldata_dic.items():
# 			print(data)
			bs_assignment_dic = {}
			for bs in self.bs_list:
				score_header = "resis_score_fraction_"+bs
# 				print(data[score_header])
				try:
					if float(data[score_header]) < threshold:
# 						print("{} = 0".format(bs))
						bs_assignment_dic[bs] = False
					elif float(data[score_header]) >= threshold:
# 						print("{} = 1".format(bs))
						bs_assignment_dic[bs] = True
				except ValueError:
					pass
# 			print(k)
			data['bs_assignments'] = bs_assignment_dic
# 			print(data)
			new_alldata_dic[k] = data
# 		print(new_alldata_dic)
		self.alldata_dic = new_alldata_dic

	def bs_lig_E_list_dic(self):
		self.alldata_assign_bs()
		dic = {}
		for bs in self.bs_list:
			dic[bs] = {}
			for lig in self.ligset_list:
				dic[bs][lig] = []
		for lig in self.ligset_list:
# 			print(lig) # check
			for k, data in self.alldata_dic.items():
				list = []
				for bs in self.bs_list:
					try:
						if (data['LIG'] == lig) and (data['bs_assignments'][bs] is True):
# 							print(data['E'])
							dic[bs][lig].append(float(data['E']))
# 						print(data)
					except KeyError:
						pass
		return dic

	def make_bs_lig_num_dic(self):
		dic = {}
		for bs, lig_E_list in self.bs_lig_E_list_dic().items():
# 			print(bs)
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
# 				print(lig)
# 				print(len(E_list))
				dic[bs][lig] = len(E_list)
		self.bs_lig_num_dic = dic


	def make_bs_lig_min_dic(self):
		dic = {}
		for bs, lig_E_list in self.bs_lig_E_list_dic().items():
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
				try:
					dic[bs][lig] = min(E_list)
				except ValueError:
					dic[bs][lig] = None
		self.bs_lig_min_dic = dic

	def make_bs_lig_avg_dic(self, threshold = 0.10 ):
		dic = {}
		for bs, lig_E_list in self.bs_lig_E_list_dic().items():
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
				try:
					dic[bs][lig] = ( sum(E_list) / len(E_list) )
				except ValueError and ZeroDivisionError:
					dic[bs][lig] = None
		self.bs_lig_avg_dic = dic

	def data_summary_by_lig(self):

		minE_dic = {}

		self.make_bs_lig_num_dic()
		self.make_bs_lig_min_dic()
		self.make_bs_lig_avg_dic()

		for lig in self.ligset_list:
			print("~~~~~~~~~~~~~~~~~~")
			print(lig)
			print("~~~~~~~~~~~~~~~~~~")

			E_list = []
			for pose, data in self.alldata_dic.items():
				if data['LIG'] == lig:
					try:
						E_list.append(float(data['E']))
					except ValueError:
						pass
# 					if data
			MinE = min(E_list)
			AvgE = mean(E_list)

			print("AvgE: {}".format(AvgE))
			print("MinE: {}".format(MinE))
# 			for alldata_head
			for bs in self.bs_list:
				print("num-{}: {}".format(bs, self.bs_lig_num_dic[bs][lig]))
				print("minE-{}: {}".format(bs, self.bs_lig_min_dic[bs][lig]))
				print("avgE-{}: {}".format(bs, self.bs_lig_avg_dic[bs][lig]))

	def data_summary_by_lig_csv(self):
		self.make_bs_lig_num_dic()
		self.make_bs_lig_min_dic()
		self.make_bs_lig_avg_dic()

		headers1 = ["Lig", "MinE", "AvgE"]
		headers2 = [] # BS counts
		headers3 = [] # BS AvgEs
		headers4 = [] # BS AvgEs

		for bs in self.bs_list:
			headers2.append("Num_"+bs) # in "+bs+" site")
			headers3.append("AvgE_"+bs)
			headers4.append("MinE_"+bs)

		headers = headers1 + headers2 + headers3 + headers4

		for h in headers:
			print(h, end=',')

		cr()

		for lig in self.ligset_list:
			print("{},{},{}".format(lig,
				min(self.alldata_lig_E_lists_dic()[lig]),
				mean(self.alldata_lig_E_lists_dic()[lig])),
				end=','
			)

# 			for bs in self.bs_list:
# 				print("{},{},{}".format("#", "avgE", "minE"),end=',')

# 			for head in ["Num_", "AvgE_", "MinE_"]:
# 				for bs in self.bs_list:
# 					print(head+bs)
			for dic in [self.bs_lig_num_dic, self.bs_lig_avg_dic, self.bs_lig_min_dic]:
				for bs in self.bs_list:
					print(dic[bs][lig],end=',')
			cr()


# 		print(len(headers))

# 		print("#adph	#fdla	#allo	avgE-adph	avgE-fdla	avgE-allo	minE-adph	minE-fdla	minE-allo")

# 		print(self.alldata_dic_sort("LIG", "aa10"))

docks_xlsx = Xlsx('/Users/zarek/lab/Docking/docks.xlsx')
basic_params = {"dock", "date", "prot", "specprot", "ligset", "box",
	"exhaust", "n_models", "n_cpus", "notes"}

def rprint(string):
	for c in str(string):
		sys.stdout.write(c)
def cr():
	print("")

def sh_list_to_py(sh_list):
	py_list = re.sub(r' ', '\", \"', sh_list)
	py_list = re.sub(r'^(.+)$', r'["\1"]', py_list)
	py_list = eval(py_list)
	return py_list

def dic_sort(dic, sortcrit, match = 0, gt_threshold = 0):
	sorted_dic = {}
	if match != 0:
		for k, v in dic.items():
			if v[sortcrit] == match:
				sorted_dic[k] = v
		return sorted_dic
	elif gt_threshold > 0:
		for k, v in dic.items():
			if v[sortcrit] >= sortval:
				sorted_dic[k] = v
		return sorted_dic

def main():
	d = Docking(dock)
# 	print(d.dock_dir)
# 	print(docks_xlsx.sheet_dic("hepi"))
# 	print(d.data_summary_by_lig())
# 	d.data_summary_by_lig()
# 	d.data_summary_by_lig_csv()
# 	print(d.bs_list)
# 	print(dock)
# 	print(d.alldata_dic_sort("LIG", "aa10"))

	test_list = "a b c d e"
# 	print(sh_list_to_py(test_list))

# 	docks_xlsx.sheet('p300')
# 	print(docks_xlsx.look_up('p300', 'p6', 'SPECPROT'))

# 	LOADING IN
# 	print(d.box_params)
	d.load_vina_results()

if __name__ == "__main__": main()
