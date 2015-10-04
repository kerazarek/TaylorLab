#!/usr/bin/env python

from openpyxl import load_workbook
from sys import argv

### Look up a value from docks.xlsx
# 	(c) 10/2/15
#
# 	use with 3 arguments
# 		1) name of sheet containing data to look up
# 		2) row to look up [the value in row 1]
# 		3) column to look up [the value in column A]

script, sheet2lookup, row2lookup, col2lookup = argv

class Xlsx():
	def __init__(self, address):
		self.address = address
		self.workbook = load_workbook(self.address)
		self.sheets_list = self.workbook.get_sheet_names()

	def sheet(self, sheet):
		return self.workbook.get_sheet_by_name(sheet)

	def look_up(self, querysheet, queryrow, querycol):
		active_sheet = self.workbook.get_sheet_by_name(querysheet)

		col_keys = {}
		for col in active_sheet.columns:
			for c in col:
				 if c.row == 1:
					col_keys[str(c.value)] = c.column

		row_keys = {}
		for col in active_sheet.columns:
			for c in col:
				 if c.column == 'A' and c.row != 1 and c.value != None:
					row_keys[str(c.value)] = c.row

		rkey  = row_keys[queryrow]
		ckey  = col_keys[querycol]
		caddress = ckey+str(rkey)
		cell = active_sheet.cell(caddress)
		contents = cell.value
		return contents

def main():
	docks_xlsx = Xlsx('/Users/zarek/lab/Docking/docks.xlsx')
# 	test = docks_xlsx.look_up('hepi', 'h2', 'EXHAUST')
# 	print(test)
	lookedup_value = docks_xlsx.look_up(sheet2lookup, row2lookup, col2lookup)
	print(lookedup_value)


if __name__ == "__main__": main()
