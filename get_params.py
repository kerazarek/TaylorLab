#!/usr/bin/env python

##################################################
###	GET PARAMETERS
##################################################
# (c) Zarek Siegel
# created 10/12/15 22:28
#
#
# 		Requires:
# 			1) docks.xlsx up to date
# 			2) pvrd results as .../$dock/pvrd_pdbqts/$dock_$lig_m$model.pdbqt
#
# 		Originally a combo of post_tsop.py and pypdb.py,
# 			eventually intended to replace pytsop
#
# 		**add a description of what it does...
#
#
# updated 10/12/15 23:37
#
#

from __future__ import print_function
import csv, re, sys
from openpyxl import load_workbook
from os.path import isfile

script, dock, print_type = sys.argv

####################
### OBJECTS
class Xlsx: # input is a .xlsx file address
	def __init__(self, address):
		self.address = address
		self.workbook = load_workbook(self.address)
		self.sheets_list = self.workbook.get_sheet_names()

	def sheet(self, sheet):
		_op = "sheet"
# 		announce(_op)

		self.active_sheet = self.workbook.get_sheet_by_name(sheet)
# 		print(self.active_sheet.columns)

		self.col_keys = {}
		self.col_heads = {}
		for col in self.active_sheet.columns:
			for c in col:
				 if c.row == 1:
					self.col_keys[str(c.value)] = c.column
					self.col_heads[c.column] = str(c.value)

		self.row_keys = {}
		self.row_heads = {}
		for col in self.active_sheet.columns:
			for c in col:
				 if c.column == 'A' and c.row != 1 and c.value != None:
					self.row_keys[str(c.value)] = c.row
					self.row_heads[c.row] = str(c.value)


	def look_up(self, querysheet, queryrow, querycol):
		_op = "look_up"
# 		announce(_op)

		self.sheet(querysheet)

		rkey = self.row_keys[queryrow]
		ckey = self.col_keys[querycol]
		caddress = ckey+str(rkey)
		cell = self.active_sheet.cell(caddress)
		contents = cell.value
		return contents


	def sheet_dic(self, sheet):
		_op = "sheet_dic"
# 		announce(_op)

		self.sheet(sheet)
		self.sheet_dic = {}
		for row in self.active_sheet.rows:
			if row[0].value != None and row[0].row != 1:
				row_head = str(row[0].value)
				row_data = {}
				for c in row:
					if c.column != 'A' and c.row != 1 and c.value != None:
						col_head = self.col_heads[c.column]
						if type(c.value) is (int or float):
							row_data[col_head] = int(c.value)
						else:
							row_data[col_head] = str(c.value)
				self.sheet_dic[row_head] = row_data
		return self.sheet_dic

class Residue: # input is a string of form 'RES123' or 'RES123_A1'
	def __init__(self, str):
# 		String
		self.str = str
# 		Atom & Residue String
		if re.search(r'^[A-Z]+[0-9]+$', self.str):
			self.atom = None
			self.res_str = self.str
		elif re.search(r'^[A-Z]+[0-9]+_.+$', self.str):
			self.atom = re.sub(r'^[A-Z]+[0-9]+_', '', self.str)
			self.res_str = re.sub(r'_[A-Z0-9]+$', '', self.str)
		else: self.atom = None
# 		Residue Index
		self.resi = re.sub(r'^[A-Z]+|_?[^_]*$', '', self.str)
		try:
			self.resi = int(self.resi)
		except ValueError:
			self.resi = None
# 		Residue Name
		self.resn = re.sub(r'[0-9]+_?[^_]*$', '', self.str)
# 		Dictionary of Props
		self.dic = {'str' : self.str, 'res_str' : self.res_str,
			'resi' : self.resi, 'resn' : self.resn, 'atom' : self.atom}

	def __str__(self):
		return self.str

class Pdb:  # input is a .pdb or .pdbqt file address which may or may not be pvr'd
	def pdb(self):
		_op = "pdb"
# 		announce(_op)

# 		print('PDB file')
		coords = []
		for line in self.pdb_lines:
			if re.search('HETATM', line) or (re.search('ATOM', line)):
				dic = {
					'atomi' : int(line[6:11]),
					'atomn' : line[12:16].replace(" ", ""),
					'resn' : line[17:20].replace(" ", ""),
					'resi' : line[22:26].replace(" ", ""),
					'x' : float(line[30:38].replace(" ", "")),
					'y' : float(line[38:46].replace(" ", "")),
					'z' : float(line[46:54].replace(" ", "")),
					'xyz' : (float(line[30:38].replace(" ", "")),
						float(line[38:46].replace(" ", "")),
						float(line[46:54].replace(" ", "")) ),
					'atom_type' : line[76:78].replace(" ", ""),
					'charge' : line[78:80].replace(" ", "") # element
				}
				coords.append(dic)
		self.coords = coords

	def pdbqt(self):
		_op = "pdbqt"
# 		announce(_op)

# 		print('PDBQT file')
		coords = []
		for line in self.pdb_lines:
# 			print(line)
			if re.search('HETATM', line) or (re.search('ATOM', line)):
				dic = {
					'atomi' : int(line[6:11]),
					'atomn' : line[12:16].replace(" ", ""),
					'resn' : line[17:20].replace(" ", ""),
					'resi' : line[22:26].replace(" ", ""),
					'x' : float(line[30:38].replace(" ", "")),
					'y' : float(line[38:46].replace(" ", "")),
					'z' : float(line[46:54].replace(" ", "")),
					'xyz' : (float(line[30:38].replace(" ", "")),
						float(line[38:46].replace(" ", "")),
						float(line[46:54].replace(" ", "")) ),
					'charge' : line[70:76].replace(" ", ""), # partial charge
					'atom_type' : line[77:79].replace(" ", "") # AD4 atom type
				}
				coords.append(dic)
		self.coords = coords

	def pvrd(self): # mine data from pvrd file
		_op = "pvrd"
# 		announce(_op)

# 		print('File has been through ADT process_VinaResult.py')
		contacts = []
		for line in self.pdb_lines:
# 			Energy
			if re.search('REMARK VINA RESULT: ', line):
				self.E = re.sub( r'^REMARK VINA RESULT:[ ]+|[ ]+[^ ]+[ ]+[^ ]+$' ,
					r'' , line.replace('\n', '')) # [23:31].replace(" ", ""))
				self.E = float(self.E)
# 			RMSD_LB
			if re.search('REMARK VINA RESULT: ', line):
				self.rmsd_lb = re.sub( r'^REMARK VINA RESULT:[ ]+[^ ]+[ ]+|[ ]+[^ ]+$' ,
					r'' , line.replace('\n', ''))
				self.rmsd_lb = float(self.rmsd_lb)
# 			RMSD_UB
			if re.search('REMARK VINA RESULT: ', line):
				self.rmsd_ub = re.sub( r'^REMARK VINA RESULT:[ ]+[^ ]+[ ]+[ ]+[^ ]+' ,
					r'' , line.replace('\n', ''))
				self.rmsd_ub = float(self.rmsd_ub)
# 			Ligand Efficiency
			if re.search('USER  AD>  ligand efficiency', line):
				self.pvr_effic = re.sub( r'USER  AD>  ligand efficiency' ,
					r'' , line.replace('\n', ''))
				self.pvr_effic = float(self.pvr_effic)
# 			Model Number
			if re.search(r'USER  AD> .+ of .+ MODELS', line):
				self.pvr_model = re.sub( r'USER  AD>| of [0-9]+ MODELS' ,
					r'' , line.replace('\n', ''))
				self.pvr_model = int(self.pvr_model)
# 			TorsDOF
			if re.search('REMARK .+ active torsions:', line):
				self.torsdof = re.sub( r'REMARK|active torsions:' ,
					r'' , line.replace('\n', ''))
				self.torsdof = int(self.torsdof)
# 			Macro_close_ats
			if re.search('USER  AD> macro_close_ats:', line):
				self.macro_close_ats = re.sub( r'USER  AD> macro_close_ats:' ,
					r'' , line.replace('\n', ''))
				self.macro_close_ats = int(self.macro_close_ats)
# 			Contacts
			if re.search(r'^USER  AD> [^ ]+:[^ ]+:[^ ]+:[^ ,]+$', line):
				contacts.append(line.replace('\n', '').replace('USER  AD> ', ''))

# 		Contacts Processing
		self.pvr_resis_objs = []
		self.pvr_resis = []
		self.pvr_resis_atoms = []
		for c in contacts:
			self.pvr_resis_objs.append(Residue(re.sub(r'^[^:]+:[^:]+:', '',
				c).replace(':', '_')))

		for r in self.pvr_resis_objs:
# 			print(r.dic)
			if r.atom != None:
				self.pvr_resis_atoms.append(r.str)
				self.pvr_resis.append(r.res_str)
			else:
				self.pvr_resis_atoms.append(None)
				self.pvr_resis.append(r.res_str)

		self.pvr_resis_objs = list(set(self.pvr_resis_objs)) # remove duplicates
		self.pvr_resis = list(set(self.pvr_resis))
		self.pvr_resis_atoms = list(set(self.pvr_resis_atoms))

		self.pvr_data = {
			'E' : self.E,
			'rmsd_ub' : self.rmsd_ub,
			'rmsd_lb' : self.rmsd_lb,
			'pvr_resis' : self.pvr_resis,
			'pvr_resis_atoms' : self.pvr_resis_atoms,
			'pvr_resis_objs' : self.pvr_resis_objs,
			'torsdof' : self.torsdof,
			'macro_close_ats' : self.macro_close_ats,
			'pvr_model' : self.pvr_model
		}

	def res_list(self, atoms = False):
		_op = "res_list"
# 		announce(_op)

		list = []
		if atoms is False:
			for atom in self.coords:
				list.append(atom['resn']+atom['resi'])
		elif atoms is True:
			for atom in self.coords:
				list.append(atom['resn']+atom['resi']+"_"+atom['atomn'])

		nodup_list = set(list)
		reslist = []
		for n in nodup_list:
			reslist.append(n)
		return reslist


	def get_type(self):
		_op = "get_type"
# 		announce(_op)

# 		Detect if its been through ADT process_VinaResult.py
		self.is_pvrd = False # by default
		for line in self.pdb_lines:
			if re.search('REMARK VINA RESULT: ', line):
				self.pvrd()
				self.is_pvrd = True

# 		Determine file type PDB/PDBQT (they are slightly different)
		if self.pdb_file_in[-5:] == 'pdbqt':
			self.pdbqt()
			self.file_type = 'pdbqt'
		elif self.pdb_file_in[-3:] == 'pdb':
			self.pdb()
			self.file_type = 'pdb'
		else:
			print("!!! BAD FILETYPE !!!")

	def __init__(self, pdb_file_in):
		self.pdb_file_in = pdb_file_in
		try:
			pdb_file_open = open(pdb_file_in)
			with pdb_file_open as f:
				self.pdb_lines = f.readlines()
		except IOError:
			pass


		self.get_type()
class BindingSite:
	def __init__(self, prot, bs):
		self.name = bs
		self.dir = prot.bs_dir
		self.lig_pdb = self.dir+"lig_pdbs/"+prot.baseprot+"_bs_"+bs+".pdb"
		self.a5res_pdb = self.dir+"a5res_pdbs/"+prot.baseprot+"_bs_"+bs+"_a5resis.pdb"

		self.lig_pdb = Pdb(self.lig_pdb)
# 		print(self.lig_pdb)
		self.a5res_pdb = Pdb(self.a5res_pdb)
		self.resis_list = self.a5res_pdb.res_list()
		self.resis_atoms_list = self.a5res_pdb.res_list(atoms = True)


class Docking: # input is a docking id e.g. a2 b34
	def basic_params(self):
		_op = "basic_params"
# 		announce(_op)

		if dock[:1] == "h":
			self.prot = "hepi"
		elif dock[:1] == "p":
			self.prot = "p300"
		else:
			print("!!! BAD DOCK ID !!!")

		self.dock = dock
		self.date = str(docks_xlsx.look_up(self.prot, dock, "DATE"))
		self.prot = str(docks_xlsx.look_up(self.prot, dock, "PROT"))
		self.specprot = str(docks_xlsx.look_up(self.prot, dock, "SPECPROT"))
		self.ligset = str(docks_xlsx.look_up(self.prot, dock, "LIGSET"))
		self.box = str(docks_xlsx.look_up(self.prot, dock, "BOX"))
		self.exhaust = int(docks_xlsx.look_up(self.prot, dock, "EXHAUST"))
		self.n_models = int(docks_xlsx.look_up(self.prot, dock, "n_MODELS"))
		self.n_cpus = docks_xlsx.look_up(self.prot, dock, "n_CPUS") # type?
		self.notes = str(docks_xlsx.look_up(self.prot, dock, "notes"))

		self.baseprot = docks_xlsx.look_up("pdbs", self.specprot, "baseprot")

		self.params = {}
		for bp in basic_params:
			evalbp = "self."+bp
			self.params[bp] = eval(evalbp)

	def ligset_params(self):
		_op = "ligset_params"
# 		announce(_op)

		self.ligset_list_sh = str(docks_xlsx.look_up("ligsets", self.ligset, "lig_list"))
		self.ligset_list = sh_list_to_py(self.ligset_list_sh)

	def box_params(self):
		_op = "box_params"
# 		announce(_op)

		self.box_params = {}
		self.box_params['name'] = str(docks_xlsx.look_up('gridboxes', self.box, 'name'))
		self.box_params['description'] = str(docks_xlsx.look_up('gridboxes', self.box, 'description'))
		self.box_params['size_x'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_x'))
		self.box_params['size_y'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_y'))
		self.box_params['size_z'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_z'))
		self.box_params['center_x'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_x'))
		self.box_params['center_y'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_y'))
		self.box_params['center_z'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_z'))
		self.box_params['notes'] = str(docks_xlsx.look_up('gridboxes', self.box, 'notes'))
		self.box_params['size_tuple'] = self.box_params['size_x'], self.box_params['size_y'], self.box_params['size_z']
		self.box_params['center_tuple'] = self.box_params['center_x'], self.box_params['center_y'], self.box_params['center_z']

		self.box_name = self.box_params['name']
		self.box_description = self.box_params['description']
		self.box_size_x = self.box_params['size_x']
		self.box_size_y = self.box_params['size_y']
		self.box_size_z = self.box_params['size_z']
		self.box_center_x = self.box_params['center_x']
		self.box_center_y = self.box_params['center_y']
		self.box_center_z = self.box_params['center_z']
		self.box_notes = self.box_params['notes']
		self.box_size_tuple = self.box_params['size_tuple']
		self.box_center_tuple = self.box_params['center_tuple']


	def binding_site_params(self):
		_op = "binding_site_params"
# 		announce(_op)

		self.bs_list_sh = docks_xlsx.look_up("pdbs", self.specprot, "binding_sites")
		self.bs_list = sh_list_to_py(self.bs_list_sh)

		if self.prot == "p300":
			self.bs_list = ["lys", "side", "coa_ado", "coa_adpp", "coa_pant", "allo1", "allo2"]

		self.bs_resis_list_dic = {}
		self.bs_resis_atoms_list_dic = {}
		for bs in self.bs_list:
			b = BindingSite(self, bs)
			self.bs_resis_list_dic[bs] = b.resis_list
			self.bs_resis_atoms_list_dic[bs] = b.resis_atoms_list
# 			eval("self."+bs+"bs_resis_list") = self.bs_resis_list_dic[bs]


	def file_addresses(self): # ALL DIRS END IN /
		_op = "file_addresses"
# 		announce(_op)

		# Basic Directories
		self.lab_dir = "/Users/zarek/lab/"
		self.docking_dir = self.lab_dir+"Docking/"
		self.prot_dir = self.docking_dir+self.prot+"/"
		self.bs_dir = self.docking_dir+"binding_sites/"+self.baseprot+"/"
		self.ligset_dir = self.docking_dir+"ligsets/"+self.ligset+"/"
		self.dock_dir = self.prot_dir+self.dock+"/"
		self.res_dir = self.dock_dir+"results/"
		self.pvrd_pdbqts_dir = self.dock_dir+"pvrd_pdbqts/"
		# Input Files
# 		self.res_pdbqt = "{}results/{}_{}_results.pdbqt".format(self.dock_dir, self.dock, self
# 		self.res_pdbqt = "{}results/{}_{}_m{}.pdbqt"
		# Output Files
		self.alldata_csv_address = self.prot_dir+self.dock+"/"+self.dock+"_alldata.csv"
		self.post_analysis_csv_address = self.prot_dir+self.dock+"/"+self.dock+"_analysis.csv"

	def fetch_params(self):
		_op = "fetch_params"
# 		announce(_op)

		self.basic_params()
		self.ligset_params()
		self.box_params()
		self.file_addresses()
		self.binding_site_params()

	def __init__(self, dock):
		self.dock = dock
		self.fetch_params()


	def print_params(self, type = None):
		if type == "readable":
			for pp in params_to_print:
				print("{}: {}".format(pp, eval("self."+pp)))
		elif type == "sh":
			for pp in params_to_print - {"bs_list"}:
				print("{}=\"{}\"".format(pp, eval("self."+pp)))
		elif type == "py":
			for pp in params_to_print - {"bs_list_sh"}:
				if eval("self."+pp) == "None":
					print("{} = {}".format(pp, None))
				elif isinstance(eval("self."+pp), float):
					print("{} = {}".format(pp, eval("self."+pp)))
				elif isinstance(eval("self."+pp), int):
					print("{} = {}".format(pp, eval("self."+pp)))
				elif isinstance(eval("self."+pp), tuple):
					print("{} = {}".format(pp, eval("self."+pp)))
				elif isinstance(eval("self."+pp), list):
					print("{} = {}".format(pp, eval("self."+pp)))
				elif isinstance(eval("self."+pp), set):
					print("{} = {}".format(pp, eval("self."+pp)))
				elif isinstance(eval("self."+pp), dict):
					print("{} = {}".format(pp, eval("self."+pp)))
				else:
					print("{} = \"{}\"".format(pp, eval("self."+pp)))
		elif type == "csv":
			for pp in params_to_print:
				print("{},{}".format(pp, eval("self."+pp)))
		else:
			print("!!! bad parameter printing type !!!")
####################

####################
# UNIVERSALS
docks_xlsx = Xlsx('/Users/zarek/lab/Docking/docks.xlsx')
basic_params = {"dock", "date", "prot", "specprot", "ligset", "box",
	"exhaust", "n_models", "n_cpus", "notes", "baseprot"}

box_params = {"box_name", "box_description", "box_size_x", "box_size_y",
	"box_size_z", "box_center_x", "box_center_y", "box_center_z",
	"box_notes", "box_size_tuple", "box_center_tuple"}

bs_params = {"bs_list", "bs_list_sh"} #, "bs_resis_list_dic", "bs_resis_atoms_list_dic"}

dir_params = {"lab_dir", "docking_dir", "prot_dir", "bs_dir", "ligset_dir",
	"dock_dir", "res_dir", "pvrd_pdbqts_dir"}

params_to_print = basic_params | box_params | bs_params | dir_params

bs_assign_threshold = 0.10
bs_assign_atom_threshold = 0.10


def cr(): print("")
def dline(): print("--------------------")
def tline(): print("~~~~~~~~~~~~~~~~~~~~")

def announce(announcement):
	cr()
	dline()
	print(announcement)
	dline()

def sh_list_to_py(sh_list):
	py_list = re.sub(r' ', '\", \"', sh_list)
	py_list = re.sub(r'^(.+)$', r'["\1"]', py_list)
	py_list = eval(py_list)
	return py_list

####################



def main():
	d = Docking(dock)
# 	d.print_params(type = "readable")
# 	tline()
# 	d.print_params(type = "sh")
# 	tline()
# 	d.print_params(type = "py")
# 	tline()
# 	d.print_params(type = "csv")
	d.print_params(type = print_type)

if __name__ == "__main__": main()
