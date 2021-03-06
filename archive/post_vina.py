#!/usr/bin/env python

##################################################
###	POST-VINA DATA PROCESSING
##################################################
# (c) Zarek Siegel
# created 10/04/15 16:57
#
#
# 		Requires:
# 			1) docks.xlsx up to date
# 			2) pvrd results as .../$dock/pvrd_pdbqts/$dock_$lig_m$model.pdbqt
#
# 		Originally a combo of post_tsop.py and pypdb.py,
# 			eventually intended to replace pytsop
#
# 		**add a description of what it does...
#
#
# updated 10/04/15 18:50
# updated 10/12/15 22:14
#
#

from __future__ import print_function
import csv, re, sys, os
from openpyxl import load_workbook
from numpy import mean
import numpy
from os.path import isfile

# KLUDGE
bs_assign_threshold = 0.10

####################
### OBJECTS
class Xlsx: # input is a .xlsx file address
	def __init__(self, address):
		self.address = address
		self.workbook = load_workbook(self.address)
		self.sheets_list = self.workbook.get_sheet_names()

	def sheet(self, sheet):
		_op = "sheet"
# 		announce(_op)

		self.active_sheet = self.workbook.get_sheet_by_name(sheet)
# 		print(self.active_sheet.columns)

		self.col_keys = {}
		self.col_heads = {}
		for col in self.active_sheet.columns:
			for c in col:
				 if c.row == 1:
					self.col_keys[str(c.value)] = c.column
					self.col_heads[c.column] = str(c.value)

		self.row_keys = {}
		self.row_heads = {}
		for col in self.active_sheet.columns:
			for c in col:
				 if c.column == 'A' and c.row != 1 and c.value != None:
					self.row_keys[str(c.value)] = c.row
					self.row_heads[c.row] = str(c.value)


	def look_up(self, querysheet, queryrow, querycol):
		_op = "look_up"
# 		announce(_op)

		self.sheet(querysheet)

		rkey = self.row_keys[queryrow]
		ckey = self.col_keys[querycol]
		caddress = ckey+str(rkey)
		cell = self.active_sheet.cell(caddress)
		contents = cell.value
		return contents


	def sheet_dic(self, sheet):
		_op = "sheet_dic"
# 		announce(_op)

		self.sheet(sheet)
		self.sheet_dic = {}
		for row in self.active_sheet.rows:
			if row[0].value != None and row[0].row != 1:
				row_head = str(row[0].value)
				row_data = {}
				for c in row:
					if c.column != 'A' and c.row != 1 and c.value != None:
						col_head = self.col_heads[c.column]
						if type(c.value) is (int or float):
							row_data[col_head] = int(c.value)
						else:
							row_data[col_head] = str(c.value)
				self.sheet_dic[row_head] = row_data
		return self.sheet_dic

class Csv: # input is a .csv file address
	def __init__(self, source_csv):
		self.source_csv = source_csv
		self.dic = {}

		csv_open = open(self.source_csv, 'r')
		csv_read = csv.reader(csv_open, delimiter=',', quotechar='"')

		csv_list = []
		for row in csv_read:
			csv_list.append(row)

		keys = csv_list[0]
		for row in csv_list:
			dic_entry = {}
			for i in range(0, len(keys)):
				if row[i] != '':
					dic_entry[keys[i]] = row[i]
			self.dic[row[0]] = dic_entry

class Residue: # input is a string of form 'RES123' or 'RES123_A1'
	def __init__(self, str):
# 		String
		self.str = str
# 		Atom & Residue String
		if re.search(r'^[A-Z]+[0-9]+$', self.str):
			self.atom = None
			self.res_str = self.str
		elif re.search(r'^[A-Z]+[0-9]+_.+$', self.str):
			self.atom = re.sub(r'^[A-Z]+[0-9]+_', '', self.str)
			self.res_str = re.sub(r'_[A-Z0-9]+$', '', self.str)
		else: self.atom = None
# 		Residue Index
		self.resi = re.sub(r'^[A-Z]+|_?[^_]*$', '', self.str)
		try:
			self.resi = int(self.resi)
		except ValueError:
			self.resi = None
# 		Residue Name
		self.resn = re.sub(r'[0-9]+_?[^_]*$', '', self.str)
# 		Dictionary of Props
		self.dic = {'str' : self.str, 'res_str' : self.res_str,
			'resi' : self.resi, 'resn' : self.resn, 'atom' : self.atom}

	def __str__(self):
		return self.str

class Pdb:  # input is a .pdb or .pdbqt file address which may or may not be pvr'd
	def pdb(self):
		_op = "pdb"
# 		announce(_op)

# 		print('PDB file')
		coords = []
		for line in self.pdb_lines:
			if re.search('HETATM', line) or (re.search('ATOM', line)):
				dic = {
					'atomi' : int(line[6:11]),
					'atomn' : line[12:16].replace(" ", ""),
					'resn' : line[17:20].replace(" ", ""),
					'resi' : line[22:26].replace(" ", ""),
					'x' : float(line[30:38].replace(" ", "")),
					'y' : float(line[38:46].replace(" ", "")),
					'z' : float(line[46:54].replace(" ", "")),
					'xyz' : (float(line[30:38].replace(" ", "")),
						float(line[38:46].replace(" ", "")),
						float(line[46:54].replace(" ", "")) ),
					'atom_type' : line[76:78].replace(" ", ""),
					'charge' : line[78:80].replace(" ", "") # element
				}
				coords.append(dic)
		self.coords = coords

	def pdbqt(self):
		_op = "pdbqt"
# 		announce(_op)

# 		print('PDBQT file')
		coords = []
		for line in self.pdb_lines:
# 			print(line)
			if re.search('HETATM', line) or (re.search('ATOM', line)):
				dic = {
					'atomi' : int(line[6:11]),
					'atomn' : line[12:16].replace(" ", ""),
					'resn' : line[17:20].replace(" ", ""),
					'resi' : line[22:26].replace(" ", ""),
					'x' : float(line[30:38].replace(" ", "")),
					'y' : float(line[38:46].replace(" ", "")),
					'z' : float(line[46:54].replace(" ", "")),
					'xyz' : (float(line[30:38].replace(" ", "")),
						float(line[38:46].replace(" ", "")),
						float(line[46:54].replace(" ", "")) ),
					'charge' : line[70:76].replace(" ", ""), # partial charge
					'atom_type' : line[77:79].replace(" ", "") # AD4 atom type
				}
				coords.append(dic)
		self.coords = coords

	def pvrd(self): # mine data from pvrd file
		_op = "pvrd"
# 		announce(_op)

# 		print('File has been through ADT process_VinaResult.py')
		contacts = []
		for line in self.pdb_lines:
# 			Energy
			if re.search('REMARK VINA RESULT: ', line):
				self.E = re.sub( r'^REMARK VINA RESULT:[ ]+|[ ]+[^ ]+[ ]+[^ ]+$' ,
					r'' , line.replace('\n', '')) # [23:31].replace(" ", ""))
				self.E = float(self.E)
# 			RMSD_LB
			if re.search('REMARK VINA RESULT: ', line):
				self.rmsd_lb = re.sub( r'^REMARK VINA RESULT:[ ]+[^ ]+[ ]+|[ ]+[^ ]+$' ,
					r'' , line.replace('\n', ''))
				self.rmsd_lb = float(self.rmsd_lb)
# 			RMSD_UB
			if re.search('REMARK VINA RESULT: ', line):
				self.rmsd_ub = re.sub( r'^REMARK VINA RESULT:[ ]+[^ ]+[ ]+[ ]+[^ ]+' ,
					r'' , line.replace('\n', ''))
				self.rmsd_ub = float(self.rmsd_ub)
# 			Ligand Efficiency
			if re.search('USER  AD>  ligand efficiency', line):
				self.pvr_effic = re.sub( r'USER  AD>  ligand efficiency' ,
					r'' , line.replace('\n', ''))
				self.pvr_effic = float(self.pvr_effic)
# 			Model Number
			if re.search(r'USER  AD> .+ of .+ MODELS', line):
				self.pvr_model = re.sub( r'USER  AD>| of [0-9]+ MODELS' ,
					r'' , line.replace('\n', ''))
				self.pvr_model = int(self.pvr_model)
# 			TorsDOF
			if re.search('REMARK .+ active torsions:', line):
				self.torsdof = re.sub( r'REMARK|active torsions:' ,
					r'' , line.replace('\n', ''))
				self.torsdof = int(self.torsdof)
# 			Macro_close_ats
			if re.search('USER  AD> macro_close_ats:', line):
				self.macro_close_ats = re.sub( r'USER  AD> macro_close_ats:' ,
					r'' , line.replace('\n', ''))
				self.macro_close_ats = int(self.macro_close_ats)
# 			Contacts
			if re.search(r'^USER  AD> [^ ]+:[^ ]+:[^ ]+:[^ ,]+$', line):
				contacts.append(line.replace('\n', '').replace('USER  AD> ', ''))

# 		Contacts Processing
		self.pvr_resis_objs = []
		self.pvr_resis = []
		self.pvr_resis_atoms = []
		for c in contacts:
			self.pvr_resis_objs.append(Residue(re.sub(r'^[^:]+:[^:]+:', '',
				c).replace(':', '_')))

		for r in self.pvr_resis_objs:
# 			print(r.dic)
			if r.atom != None:
				self.pvr_resis_atoms.append(r.str)
				self.pvr_resis.append(r.res_str)
			else:
				self.pvr_resis_atoms.append(None)
				self.pvr_resis.append(r.res_str)

		self.pvr_resis_objs = list(set(self.pvr_resis_objs)) # remove duplicates
		self.pvr_resis = list(set(self.pvr_resis))
		self.pvr_resis_atoms = list(set(self.pvr_resis_atoms))

		self.pvr_data = {
			'E' : self.E,
			'rmsd_ub' : self.rmsd_ub,
			'rmsd_lb' : self.rmsd_lb,
			'pvr_resis' : self.pvr_resis,
			'pvr_resis_atoms' : self.pvr_resis_atoms,
			'pvr_resis_objs' : self.pvr_resis_objs,
			'torsdof' : self.torsdof,
			'macro_close_ats' : self.macro_close_ats,
			'pvr_model' : self.pvr_model
		}

	def res_list(self, atoms = False):
		_op = "res_list"
# 		announce(_op)

		list = []
		if atoms is False:
			for atom in self.coords:
				list.append(atom['resn']+atom['resi'])
		elif atoms is True:
			for atom in self.coords:
				list.append(atom['resn']+atom['resi']+"_"+atom['atomn'])

		nodup_list = set(list)
		reslist = []
		for n in nodup_list:
			reslist.append(n)
		return reslist


	def get_type(self):
		_op = "get_type"
# 		announce(_op)

# 		Detect if its been through ADT process_VinaResult.py
		self.is_pvrd = False # by default
		for line in self.pdb_lines:
			if re.search('REMARK VINA RESULT: ', line):
				self.pvrd()
				self.is_pvrd = True

# 		Determine file type PDB/PDBQT (they are slightly different)
		if self.pdb_file_in[-5:] == 'pdbqt':
			self.pdbqt()
			self.file_type = 'pdbqt'
		elif self.pdb_file_in[-3:] == 'pdb':
			self.pdb()
			self.file_type = 'pdb'
		else:
			print("!!! BAD FILETYPE !!!")

	def __init__(self, pdb_file_in):
		self.pdb_file_in = pdb_file_in
		try:
			pdb_file_open = open(pdb_file_in)
			with pdb_file_open as f:
				self.pdb_lines = f.readlines()
		except IOError:
			pass


		self.get_type()

class Pose:
	def __init__(self, key = None, dock = None, lig = None, model = None):
		if key == None and dock == None and lig == None and model == None:
			print("Pose needs a specified key, or lig/model")
		elif key != None and dock == None and lig == None and model == None:
			if re.search(r'^[a-z][0-9]+_?[a-z0-9]*_[A-Za-z0-9_]+_m[0-9]+$', key):
				self.key = key
				if re.search(r'^[a-z][0-9]+_[A-Za-z0-9_]+_m[0-9]+$', key):
					self.dock = re.sub(r'_[A-Za-z0-9_]+_m[0-9]+$', '', key)
					self.lig = re.sub(r'^[a-z][0-9]+_|_m[0-9]+$', '', key)
					self.model = int(re.sub(r'^[a-z][0-9]+_[A-Za-z0-9_]+_m', '', key))
			else: print("!!! bad pose key !!!")
		elif key == None and dock != None and lig != None and model != None:
			self.dock = dock
			self.lig = lig
			self.model = model
			self.key = dock+"_"+lig+"_m"+str(model)
		else: print("!!! ERROR GENERATING POSE !!!")

	def pdb_import(self, docking_object):
		#pvrd_pdbqt
		self.pvrd_pdbqt = docking_object.pvrd_pdbqts_dir+self.key+".pdbqt"
		self.pdb = Pdb(self.pvrd_pdbqt)


class BindingSite:
	def __init__(self, prot, bs):
		self.name = bs
		self.dir = prot.bs_dir
		self.lig_pdb = self.dir+"lig_pdbs/"+prot.baseprot+"_bs_"+bs+".pdb"
		self.a5res_pdb = self.dir+"a5res_pdbs/"+prot.baseprot+"_bs_"+bs+"_a5resis.pdb"

		self.lig_pdb = Pdb(self.lig_pdb)
# 		print(self.lig_pdb)
		self.a5res_pdb = Pdb(self.a5res_pdb)
		self.resis_list = self.a5res_pdb.res_list()
		self.resis_atoms_list = self.a5res_pdb.res_list(atoms = True)

# 		print(self.name, self.dir, self.lig_pdb, self.a5res_pdb)
# 		print(self.resis_list, self.resis_atoms_list)

class Docking: # input is a docking id e.g. a2 b34
	def basic_params(self):
		_op = "basic_params"
# 		announce(_op)

		if dock[:1] == "h":
			self.prot = "hepi"
		elif dock[:1] == "p":
			self.prot = "p300"
		else:
			print("!!! BAD DOCK ID !!!")

		self.dock = dock
		self.date = str(docks_xlsx.look_up(self.prot, dock, "DATE"))
		self.prot = str(docks_xlsx.look_up(self.prot, dock, "PROT"))
		self.specprot = str(docks_xlsx.look_up(self.prot, dock, "SPECPROT"))
		self.ligset = str(docks_xlsx.look_up(self.prot, dock, "LIGSET"))
		self.box = str(docks_xlsx.look_up(self.prot, dock, "BOX"))
		self.exhaust = str(docks_xlsx.look_up(self.prot, dock, "EXHAUST"))
		self.n_models = int(docks_xlsx.look_up(self.prot, dock, "n_MODELS"))
		self.n_cpus = docks_xlsx.look_up(self.prot, dock, "n_CPUS") # type?
		self.notes = str(docks_xlsx.look_up(self.prot, dock, "notes"))

		self.baseprot = docks_xlsx.look_up("pdbs", self.specprot, "baseprot")

		self.params = {}
		for bp in basic_params:
			evalbp = "self."+bp
			self.params[bp] = eval(evalbp)

	def ligset_params(self):
		_op = "ligset_params"
# 		announce(_op)

		self.ligset_list_sh = str(docks_xlsx.look_up("ligsets", self.ligset, "lig_list"))
		self.ligset_list = sh_list_to_py(self.ligset_list_sh)

	def box_params(self):
		_op = "box_params"
# 		announce(_op)

		self.box_params = {}
		self.box_params['name'] = str(docks_xlsx.look_up('gridboxes', self.box, 'name'))
		self.box_params['description'] = str(docks_xlsx.look_up('gridboxes', self.box, 'description'))
		self.box_params['size_x'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_x'))
		self.box_params['size_y'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_y'))
		self.box_params['size_z'] = float(docks_xlsx.look_up('gridboxes', self.box, 'size_z'))
		self.box_params['center_x'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_x'))
		self.box_params['center_y'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_y'))
		self.box_params['center_z'] = float(docks_xlsx.look_up('gridboxes', self.box, 'center_z'))
		self.box_params['notes'] = str(docks_xlsx.look_up('gridboxes', self.box, 'notes'))
		self.box_params['size_tuple'] = self.box_params['size_x'], self.box_params['size_y'], self.box_params['size_z']
		self.box_params['center_tuple'] = self.box_params['center_x'], self.box_params['center_y'], self.box_params['center_z']
		self.box_params['name'] = str(docks_xlsx.look_up('gridboxes', self.box, 'name'))

		self.box_center_x = self.box_params['center_x']
		self.box_center_y = self.box_params['center_y']
		self.box_center_z = self.box_params['center_z']
		self.box_size_x = self.box_params['size_x']
		self.box_size_y = self.box_params['size_y']
		self.box_size_z = self.box_params['size_z']

	def binding_site_params(self):
		_op = "binding_site_params"
# 		announce(_op)

		self.bs_list_sh = docks_xlsx.look_up("pdbs", self.specprot, "binding_sites")
		self.bs_list = sh_list_to_py(self.bs_list_sh)

		if self.prot == "p300":
			self.bs_list = ["lys", "side", "coa", "allo1", "allo2", "coa_ado", "coa_adpp", "coa_pant"] # ["lys", "side", "coa", "coa_ado", "coa_adpp", "coa_pant", "allo1", "allo2"]

		self.bs_resis_list_dic = {}
		self.bs_resis_atoms_list_dic = {}
		for bs in self.bs_list:
			b = BindingSite(self, bs)
			self.bs_resis_list_dic[bs] = b.resis_list
			self.bs_resis_atoms_list_dic[bs] = b.resis_atoms_list


	def file_addresses(self): # ALL DIRS END IN /
		_op = "file_addresses"
# 		announce(_op)

		# Basic Directories
		self.lab_dir = "/Users/zarek/lab/"
		self.docking_dir = self.lab_dir+"Docking/"
		self.prot_dir = self.docking_dir+self.prot+"/"
		self.bs_dir = self.docking_dir+"binding_sites/"+self.baseprot+"/"
		self.ligset_dir = self.docking_dir+"ligsets/"+self.ligset+"/"
		self.dock_dir = self.prot_dir+self.dock+"/"
		self.res_dir = self.dock_dir+"results/"
		self.pvrd_pdbqts_dir = self.dock_dir+"pvrd_pdbqts/"

		self.cluster_home = "/home/zsiegel/"
		# Input Files
# 		self.res_pdbqt = "{}results/{}_{}_results.pdbqt".format(self.dock_dir, self.dock, self
# 		self.res_pdbqt = "{}results/{}_{}_m{}.pdbqt"
		# Output Files
		self.alldata_csv_address = self.prot_dir+self.dock+"/"+self.dock+"_alldata.csv"
		self.post_analysis_csv_address = self.prot_dir+self.dock+"/"+self.dock+"_analysis.csv"

		self.specprot_pdbqt = self.cluster_home+self.prot+"/"+self.specprot+".pdbqt"
	def fetch_params(self):
		_op = "fetch_params"
#		announce(_op)

		self.basic_params()
		self.ligset_params()
		self.box_params()
		self.file_addresses()
		self.binding_site_params()

	def init_meta(self):
		self.missing_keys = []
		self.is_alldata_dic_generated = False

	def __init__(self, dock):
		self.dock = dock
		self.fetch_params()
		self.init_meta()

	def load_vina_results(self):
		_op = "load_vina_results"
#		announce(_op)

		for lig in self.ligset_list:
			print(lig)
			res_pdbqt = self.res_dir+self.dock+"_"+lig+"_results.pdbqt"
			for m in range(1, (int(self.n_models) + 1)):
				try:
					pvrd_pdbqt = str(self.pvrd_pdbqts_dir)+self.dock+"_"+lig+"_m"+str(m)+".pdbqt"
					pvrd = Pvrd(pvrd_pdbqt)
# 					print(pvrd.dic)
				except IOError:
					pass

# 	def write_bsub(self):
# 		_output_template = "#BSUB -q hp12\n\
# #BSUB -n {n_cpus}\n\
# #BSUB -o {outlog}\n\
# #BSUB -e {errlog}\n\
# #BSUB -J {jobname}\n\n\
# echo \"~~~BEGIN DOCKING {dock}~~~\"\n\n\
# command mkdir {dock}\n\
# command mkdir {dock}/results\n\n\
# for lig in {ligset_list}\n\
# do\n\
# 	/share/apps/autodock/autodock_vina_1_1_2_linux_x86/bin/vina \\\n\
# 	--receptor {specprot_pdbqt} \\\n\
# 	--ligand {lig_pdbqt} \\\n\
# 	--out {res_pdbqt} \\\n\
# 	--center_x {box_center_x} \\\n\
# 	--center_y {box_center_y} \\\n\
# 	--center_z {box_center_z} \\\n\
# 	--size_x {box_size_x} \\\n\
# 	--size_y {box_size_y} \\\n\
# 	--size_z {box_size_z} \\\n\
# 	--cpu {n_cpus} \\\n\
# 	--num_modes {n_models} \\\n\
# 	--exhaustiveness {exhaust}\n\
# 	echo finished docking $lig of docking {dock}\n\
# done\n\n\
# echo \"~~~END DOCKING {dock}~~~\""
#
# 		if self.n_models <= 20:
# 			_n_models = self.n_models
# 			_dock = self.dock
# 			if type(self.n_cpus) is not int: self.n_cpus = 1
# # 			_output_address = self.prot_dir+"vsub_"+_dock
# 			_output_address = self.dock_dir+"vsub_"+_dock
# # 			print(_output_address)
# 			output = (_output_template.format(
# 				n_cpus = self.n_cpus,
# 				outlog = "dock_logs/"+_dock+"_log_out.txt",
# 				errlog = "dock_logs/"+_dock+"_log_err.txt",
# 				jobname = "VinaDock_"+_dock,
# 				dock = _dock,
# 				ligset_list = self.ligset_list_sh,
# 				specprot_pdbqt = self.specprot_pdbqt,
# 				lig_pdbqt = self.cluster_home+"ligsets/"+self.ligset+"/$lig.pdbqt",
# 				res_pdbqt = self.cluster_home+_dock+"/results/"+_dock+"_$lig\_results.pdbqt",
# 				box_center_x = self.box_params['center_x'],
# 				box_center_y = self.box_params['center_y'],
# 				box_center_z = self.box_params['center_z'],
# 				box_size_x = self.box_params['size_x'],
# 				box_size_y = self.box_params['size_y'],
# 				box_size_z = self.box_params['size_z'],
# 				n_models = _n_models,
# 				exhaust = self.exhaust
# 			))
# 			f = open(_output_address, 'w')
# 			f.write(str(output))
# # 			print("Vina submission script is at: {}".format(_output_address))
# 		elif self.n_models > 20:
# 			if (self.n_models % 20) != 0:
# 				print("!!! if more than 20 models, gotta be multiple of 20 !!!")
#
# 			_n_subdocks = self.n_models / 20
# 			for subdock in range(1, _n_subdocks + 1):
# 				_n_models = 20
# 				_dock = self.dock+"."+str(subdock)
# # 				_output_address = self.prot_dir+"vsub_"+_dock
# 				_output_address = self.dock_dir+"vsub_"+_dock
# 				output = (_output_template.format(
# 					n_cpus = self.n_cpus,
# 					outlog = "dock_logs/"+_dock+"_log_out.txt",
# 					errlog = "dock_logs/"+_dock+"_log_err.txt",
# 					jobname = "VinaDock_"+_dock,
# 					dock = _dock,
# 					ligset_list = self.ligset_list_sh,
# 					specprot_pdbqt = self.specprot_pdbqt,
# 					lig_pdbqt = self.cluster_home+"ligsets/"+self.ligset+"/$lig.pdbqt",
# 					res_pdbqt = self.cluster_home+_dock+"/results/"+_dock+"_$lig\_results.pdbqt",
# 					box_center_x = self.box_params['center_x'],
# 					box_center_y = self.box_params['center_y'],
# 					box_center_z = self.box_params['center_z'],
# 					box_size_x = self.box_params['size_x'],
# 					box_size_y = self.box_params['size_y'],
# 					box_size_z = self.box_params['size_z'],
# 					n_models = _n_models,
# 					exhaust = self.exhaust
# 				))
# 				f = open(_output_address, 'w')
# 				f.write(str(output))

	def write_bsub(self, multi=False):
		_output_template = "#BSUB -q hp12\n\
#BSUB -n {n_cpus}\n\
#BSUB -o {outlog}\n\
#BSUB -e {errlog}\n\
#BSUB -J {jobname}\n\n\
echo \"~~~BEGIN DOCKING {dock}~~~\"\n\n\
command mkdir {dock_dir}results\n\n\
for lig in {ligset_list}\n\
do\n\
	/share/apps/autodock/autodock_vina_1_1_2_linux_x86/bin/vina \\\n\
	--receptor {specprot_pdbqt} \\\n\
	--ligand {lig_pdbqt} \\\n\
	--out {dock_dir}results/{dock}_$lig\_results.pdbqt \\\n\
	--center_x {box_center_x} \\\n\
	--center_y {box_center_y} \\\n\
	--center_z {box_center_z} \\\n\
	--size_x {box_size_x} \\\n\
	--size_y {box_size_y} \\\n\
	--size_z {box_size_z} \\\n\
	--cpu {n_cpus} \\\n\
	--num_modes {n_models} \\\n\
	--exhaustiveness {exhaust}\n\
	echo finished docking $lig of docking {dock}\n\
done\n\n\
echo \"~~~END DOCKING {dock}~~~\""

		_n_models = self.n_models
		if multi:
			_base_dock = re.sub(r'\..+$', '', self.dock)
# 			_res_pdbqt = self.cluster_home+_base_dock+"/"+self.dock+"/results/"+self.dock+"_$lig\_results.pdbqt"
			_dock_dir = self.cluster_home+_base_dock+"/"+self.dock+"/"
		else:
# 			_res_pdbqt = self.cluster_home+self.dock+"/results/"+self.dock+"_$lig\_results.pdbqt"
			_dock_dir = self.cluster_home+self.dock+"/"

		if type(self.n_cpus) is not int: self.n_cpus = 1
# 			_output_address = self.prot_dir+"vsub_"+self.dock
		_output_address = self.dock_dir+"vsub_"+self.dock
# 			print(_output_address)
		output = (_output_template.format(
			n_cpus = self.n_cpus,
			outlog = "dock_logs/"+self.dock+"_log_out.txt",
			errlog = "dock_logs/"+self.dock+"_log_err.txt",
			jobname = "VinaDock_"+self.dock,
			dock = self.dock,
			dock_dir = _dock_dir,
			ligset_list = self.ligset_list_sh,
			specprot_pdbqt = self.specprot_pdbqt,
			lig_pdbqt = self.cluster_home+"ligsets/"+self.ligset+"/$lig.pdbqt",
			box_center_x = self.box_params['center_x'],
			box_center_y = self.box_params['center_y'],
			box_center_z = self.box_params['center_z'],
			box_size_x = self.box_params['size_x'],
			box_size_y = self.box_params['size_y'],
			box_size_z = self.box_params['size_z'],
			n_models = self.n_models,
			exhaust = self.exhaust
		))
		f = open(_output_address, 'w')
		f.write(str(output))

	def write_params(self):
		for p in print_params:
# 			ep = "print(self.{}, end='')".format(p)
# 			print("{}=\"".format(p), end='')
# 			eval(ep)
# 			print("\"")
# 			print(p)
			pass


		output="baseprot=\"{baseprot}\"\n\
box=\"{box}\"\n\
box_center_x=\"{box_center_x}\"\n\
box_center_y=\"{box_center_y}\"\n\
box_center_z=\"{box_center_z}\"\n\
box_size_x=\"{box_size_x}\"\n\
box_size_y=\"{box_size_y}\"\n\
box_size_z=\"{box_size_z}\"\n\
date=\"{date}\"\n\
dock=\"{dock}\"\n\
exhaust=\"{exhaust}\"\n\
ligset=\"{ligset}\"\n\
ligset_list_sh=\"{ligset_list_sh}\"\n\
n_cpus=\"{n_cpus}\"\n\
n_models=\"{n_models}\"\n\
notes=\"{notes}\"\n\
prot=\"{prot}\"\n\
specprot=\"{specprot}\"\n\
specprot_pdbqt=\"{specprot_pdbqt}\"".format(
	baseprot=self.baseprot, box=self.box, box_center_x=self.box_center_x,
	box_center_y=self.box_center_y, box_center_z=self.box_center_z,
	box_size_x=self.box_size_x, box_size_y=self.box_size_y,
	box_size_z=self.box_size_z, date=self.date, dock=self.dock,
	exhaust=self.exhaust, ligset=self.ligset, ligset_list_sh=self.ligset_list_sh,
	n_cpus=self.n_cpus, n_models=self.n_models, notes=self.notes, prot=self.prot,
	specprot=self.specprot, specprot_pdbqt=self.specprot_pdbqt
	)

# 		_output_address = self.prot_dir+dock+"_params.txt"
		_output_address = self.dock_dir+dock+"_params.txt"
		f = open(_output_address, 'w')
		f.write(str(output))
# 		print("Params txt is at: {}".format(_output_address))

	def pre(self, pre_name):
		self.dock_dir = self.prot_dir+pre_name+"/"
		os.mkdir(self.dock_dir)
		self.write_params()

		if self.n_models <= 20:
			self.write_bsub()
		else:
			if (self.n_models % 20) != 0:
					print("!!! if more than 20 models, gotta be multiple of 20 !!!")
			else: self.n_subdocks = self.n_models / 20

			_dock_dir = self.dock_dir
			for s in range(1, self.n_subdocks + 1):
				sd = d
				sd.dock = pre_name+"."+str(s)
				sd.n_models = 20
				sd.dock_dir = _dock_dir+sd.dock+"/"
				os.mkdir(sd.dock_dir)
				sd.write_params()
				sd.write_bsub(multi=True)



# 	print("Dock folder (with vsub and params) is at: {}".format(self.dock_dir))



	def generate_alldata_dic(self):
		_op = "generate_alldata_dic"
#		announce(_op)

		self.alldata_dic = {}
		self.bs_key_list_dic = {}

		for bs in self.bs_list:
			self.bs_key_list_dic[bs] = []

		for l in self.ligset_list:
# 			print("~~~[{}]~~~".format(l))

			for m in range(1, (self.n_models + 1)):
# 				print("[{}]".format(m))
				pose = Pose(dock = self.dock, lig = l, model = m)
				self.alldata_dic[pose.key] = {}

				pvrd_pdbqt = self.pvrd_pdbqts_dir+pose.key+".pdbqt"
				if isfile(pvrd_pdbqt):
					pvrd_obj = Pdb(pvrd_pdbqt)
	# 				print(pvrd_pdbqt)

					_lm_lig = l
					_lm_model = m
					_lm_key = pose.key

					_lm_E = pvrd_obj.E
					_lm_pvr_resis = set(pvrd_obj.pvr_resis)
					_lm_pvr_resis_atoms = set(pvrd_obj.pvr_resis_atoms)
					_lm_rmsd_ub = pvrd_obj.rmsd_ub
					_lm_rmsd_lb = pvrd_obj.rmsd_lb
					_lm_pvr_effic = pvrd_obj.pvr_effic
					_lm_pvr_model = pvrd_obj.pvr_model
					_lm_macro_close_ats = pvrd_obj.macro_close_ats
					_lm_torsdof = pvrd_obj.torsdof

# 					if m != _lm_pvr_model: print("!!! pvr model mismatch <{}_m{}> !!! ([pvr_model: {}; iter_model: {})".format(l, m, _lm_pvr_model, m))

					_lm_bs_resis_counts = {}
					_lm_bs_resis_atoms_counts = {}
					_lm_bs_resis_fractions = {}
					_lm_bs_resis_atoms_fractions = {}
					_lm_bs_assignments = [] # resis
					_lm_bs_atoms_assignments = [] # resis atoms

# 					_lm_bs_resis_counts['total'] = 0
# 					_lm_bs_resis_atoms_counts['total'] = 0
					for bs in self.bs_list:
						_lm_bs_resis_counts[bs] = 0
						_lm_bs_resis_atoms_counts[bs] = 0
						for r in self.bs_resis_list_dic[bs]:
							if r in _lm_pvr_resis:
								_lm_bs_resis_counts[bs] += 1
# 								_lm_bs_resis_counts['total'] += 1
						for r in self.bs_resis_atoms_list_dic[bs]:
							if r in _lm_pvr_resis_atoms:
								_lm_bs_resis_atoms_counts[bs] += 1
# 								_lm_bs_resis_atoms_counts['total'] += 1

						_lm_bs_resis_fractions[bs] = r4(float(_lm_bs_resis_counts[bs]) / len(self.bs_resis_list_dic[bs]))
						_lm_bs_resis_atoms_fractions[bs] = r4(float(_lm_bs_resis_atoms_counts[bs]) / len(self.bs_resis_atoms_list_dic[bs]))

						if _lm_bs_resis_fractions[bs] >= bs_assign_threshold:
							_lm_bs_assignments.append(bs)
							self.bs_key_list_dic[bs].append(pose.key)

						if _lm_bs_resis_atoms_fractions[bs] >= bs_assign_atom_threshold:
							_lm_bs_atoms_assignments.append(bs)

					_lm_bs_assignments = set(_lm_bs_assignments)
					_lm_bs_atoms_assignments = set(_lm_bs_atoms_assignments)

					params = pvr_params | pose_params | bs_params
					for pp in params:
						_lm_pp = "_lm_"+pp
						self.alldata_dic[pose.key][pp] = eval(_lm_pp)
						none_pp = "{} = None".format(_lm_pp)
						exec none_pp

				else:
# 					print("Ligand {} m{} does pvrd_pdbqt does not exist".format(l, m))
					pass

		self.is_alldata_dic_generated = True

	def __iter__(self):
		if not self.is_alldata_dic_generated:
			self.generate_alldata_dic()
		for key, data in self.alldata_dic.items():
			yield data

	def alldata_lig_subsets_dic(self):
		_op = "alldata_lig_subsets_dic"
#		announce(_op)

		dic = {}
		for lig in self.ligset_list:
			dic[lig] = dic_subset(self.alldata_dic, "lig", subset_value = lig)
		return dic

# 	is this one even useful?
	def generate_alldata_lig_E_lists_dic(self):
		_op = "alldata_lig_E_lists_dic"
#		announce(_op)

		dic = {}
		for lig, subset_dic in self.alldata_lig_subsets_dic().items():
			lig_E_list = []
			for k, v in subset_dic.items():
				lig_E_list.append(float(v['E']))
			dic[lig] = lig_E_list
		self.alldata_lig_E_lists_dic = dic

	def make_bs_lig_E_list_dic(self):
		_op = "make_bs_lig_E_list_dic"
#		announce(_op)

		_dic = {}

		for bs in self.bs_list:
			_dic[bs] = {}
			for lig in self.ligset_list:
# 				print(lig)
				_dic[bs][lig] = []
		for key, data in self.alldata_dic.items():
			try: # KLUDGE
				for assigned_bs in data['bs_assignments']:
					if type(data['E']) is float:
						_dic[assigned_bs][data['lig']].append(data['E'])
			except KeyError: pass

# 		print(_dic)
		self.bs_lig_E_list_dic = _dic

	def make_bs_lig_num_dic(self):
		_op = "make_bs_lig_num_dic"
#		announce(_op)

		dic = {}

		for bs, lig_E_list in self.bs_lig_E_list_dic.items():
# 			print(bs)
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
# 				print(lig)
# 				print(len(E_list))
				dic[bs][lig] = len(E_list)
		self.bs_lig_num_dic = dic


	def make_bs_lig_min_dic(self):
		_op = "make_bs_lig_min_dic"
#		announce(_op)

		dic = {}
		for bs, lig_E_list in self.bs_lig_E_list_dic.items():
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
				try:
					dic[bs][lig] = min(E_list)
				except ValueError:
					dic[bs][lig] = None
		self.bs_lig_min_dic = dic

	def make_bs_lig_avg_dic(self, threshold = bs_assign_threshold ):
		_op = "make_bs_lig_avg_dic"
#		announce(_op)

		dic = {}
		for bs, lig_E_list in self.bs_lig_E_list_dic.items():
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
				try:
					E_array = numpy.array(E_list)
					if numpy.isnan(numpy.mean(E_array, axis=0)):
						dic[bs][lig] = None
					else:
						dic[bs][lig] = numpy.mean(E_array, axis=0)
				except: # ValueError and ZeroDivisionError:
					dic[bs][lig] = None
		self.bs_lig_avg_dic = dic

	def make_bs_lig_stdev_dic(self, threshold = bs_assign_threshold ):
		_op = "make_bs_lig_avg_dic"
#		announce(_op)

		dic = {}
		for bs, lig_E_list in self.bs_lig_E_list_dic.items():
			dic[bs] = {}
			for lig, E_list in lig_E_list.items():
				try:
					E_array = numpy.array(E_list)
					if numpy.isnan(numpy.std(E_array, axis=0)):
						dic[bs][lig] = None
					else:
						dic[bs][lig] = numpy.std(E_array, axis=0)
				except:
					dic[bs][lig] = None
		self.bs_lig_stdev_dic = dic

	def data_summary_by_lig(self):
		_op = "data_summary_by_lig"
#		announce(_op)

		minE_dic = {}

		self.make_bs_lig_E_list_dic()

		self.make_bs_lig_num_dic()
		self.make_bs_lig_min_dic()
		self.make_bs_lig_avg_dic()
		self.make_bs_lig_stdev_dic()

		for lig in self.ligset_list:
			cr()
			tline()
			print(lig)
			tline()

			E_list = []
			for pose, data in self.alldata_dic.items():
				if data['lig'] == lig:
					try:
						E_list.append(float(data['E']))
					except ValueError:
						pass
# 					if data
			MinE = min(E_list)
			AvgE = mean(E_list)
			StdevE = numpy.std(numpy.array(E_list), axis=0)


			print("AvgE: {}".format(AvgE))
			print("MinE: {}".format(MinE))
# 			for alldata_head
			for bs in self.bs_list:
				print("num-{}: {}".format(bs, self.bs_lig_num_dic[bs][lig]))
				print("minE-{}: {}".format(bs, self.bs_lig_min_dic[bs][lig]))
				print("avgE-{}: {}".format(bs, self.bs_lig_avg_dic[bs][lig]))

	def data_summary_by_lig_csv(self):
		_op = "data_summary_by_lig_csv"
#		announce(_op)

		self.make_bs_lig_E_list_dic()
		self.make_bs_lig_num_dic()
		self.make_bs_lig_min_dic()
		self.make_bs_lig_avg_dic()
		self.make_bs_lig_stdev_dic()

		self.generate_alldata_lig_E_lists_dic()

		headers1 = ["Lig", "MinE", "AvgE", "StdevE"]
		headers2 = [] # BS counts
		headers3 = [] # BS AvgEs
		headers4 = [] # BS AvgEs
		headers5 = []

		for bs in self.bs_list:
			headers2.append("Num_"+bs) # in "+bs+" site")
			headers3.append("AvgE_"+bs)
			headers4.append("MinE_"+bs)
			headers4.append("StdevE_"+bs)

		headers = headers1 + headers2 + headers3 + headers4 + headers5

		for h in headers:
			print(h, end=',')

		cr()



		for lig in self.ligset_list:
			print("{},{},{},{}".format(lig,
				r4(min(self.alldata_lig_E_lists_dic[lig])),
				r4(mean(self.alldata_lig_E_lists_dic[lig])),
				r4(numpy.std(numpy.array(self.alldata_lig_E_lists_dic[lig]), axis=0))),
				end=','
			)

# 			for bs in self.bs_list:
# 				print("{},{},{}".format("#", "avgE", "minE"),end=',')

# 			for head in ["Num_", "AvgE_", "MinE_"]:
# 				for bs in self.bs_list:
# 					print(head+bs)
			for dic in [self.bs_lig_num_dic, self.bs_lig_avg_dic, self.bs_lig_min_dic, self.bs_lig_stdev_dic]:
				for bs in self.bs_list:
					try:
						if type(dic[bs][lig]) is int:
							print(dic[bs][lig],end=',')
						else:
							print(r4(dic[bs][lig]),end=',')
					except TypeError:
						print('',end=',')
			cr()

	def print_alldata_csv(self):
		keys = [
			"key", "lig", "model", "E", "pvr_resis", "pvr_resis_atoms"
		]
		print(py_list_to_csv(keys))
		for data in self:
			csv_row = []
			for k in keys:
# 				if type(data[k]) is (set or list): print(py_list_to_sh(data[k]))
# 				else: print(data[k])
				try:
					if type(data[k]) is (set or list):
						csv_row.append((py_list_to_sh(data[k])))
					else: csv_row.append(data[k])
				except:
					csv_row.append(None)
			print(py_list_to_csv(csv_row))

	def generate_allposes_dic(self): # incorporate into alldata dic?
		self.allposes_dic = {}
		for data in self:
			pose = Pose(data['key'])
			pose.pdb_import(self)
			self.allposes_dic[data['key']] = pose.pdb




####################

####################
# UNIVERSALS
docks_xlsx = Xlsx('/Users/zarek/lab/Docking/docks.xlsx')
basic_params = {"dock", "date", "prot", "specprot", "ligset", "box",
	"exhaust", "n_models", "n_cpus", "notes", "baseprot"}
print_params = basic_params | {"ligset_list_sh", "specprot_pdbqt",
	"box_center_x", "box_center_y", "box_center_z",
	"box_size_x", "box_size_y", "box_size_z"}
pose_params = {"lig", "model", "key"}
pvr_params = {"E", "pvr_resis", "pvr_resis_atoms", "rmsd_ub", "rmsd_lb",
	"pvr_effic", "pvr_model", "torsdof", "macro_close_ats"}
bs_params = {"bs_resis_counts", "bs_resis_atoms_counts",
	"bs_resis_fractions", "bs_resis_atoms_fractions",
	"bs_assignments", "bs_atoms_assignments"}
# bs_assign_threshold = 0.01 # KLUDGE
# bs_assign_atom_threshold = 0.10
bs_assign_atom_threshold = bs_assign_threshold

def cr(): print("")
def dline(): print("--------------------")
def tline(): print("~~~~~~~~~~~~~~~~~~~~")

def announce(announcement):
# 	cr()
	dline()
	print(announcement)
	dline()

def r4(x): return round(x, 4)

def sh_list_to_py(sh_list):
	py_list = re.sub(r' ', '\",\"', sh_list)
	py_list = re.sub(r'^(.+)$', r'["\1"]', py_list)
	py_list = eval(py_list)
	return py_list

def py_list_to_sh(py_list):
	if type(py_list) is set: py_list = list(py_list)
	list_str = str(py_list)
	sh_list = re.sub(r'[,\'\[\]]', '', list_str)
	return sh_list

def py_list_to_csv(py_list):
	list_str = str(py_list)
	csv_list = re.sub(r'[\'\[\]]', '', list_str)
	csv_list = re.sub(r', ', ',', csv_list)
	return csv_list

def dic_subset(dic, subset_key, subset_value = 0, gt_threshold = 0):
	subsetted_dic = {}
	if subset_value != 0:
		for k, v in dic.items():
			if not (v == {}):
				if v[subset_key] == subset_value:
					subsetted_dic[k] = v
		return subsetted_dic
	elif gt_threshold > 0:
		for k, v in dic.items():
			if not (v == {}):
				if v[subset_key] >= gt_threshold:
					subsetted_dic[k] = v
		return subsetted_dic
	else: print('subsetting requires proper criteria')

def n2b(x): # none -> blank
	if x is not None:
		return x
	else:
		return ''


####################


def main():
	announce_start = "START"
# 	announce(announce_start)

	def execute_help():
		print("\n	HELP:\n\
		  * Required: at least 1 argument (dock id)\n")
	def initiate():
		print("initiate")
		global dock, d
		dock = sys.argv[1]
		d = Docking(dock)
	# 	d = Docking(dock)
	def execute_default():
		print("no arguments given (default)")
	def execute_pre():
		announce('pre')
		if len(sys.argv) > 3:
			pre_name = sys.argv[3]
		else: pre_name = dock
		d.pre(pre_name)


	if len(sys.argv) < 2:
		announce("!!! not enough arguments !!!")
		execute_help()
	elif len(sys.argv) == 2:
		if (sys.argv[1] == 'h') or (sys.argv[1] == 'help'):
			execute_help()
		else:
			initiate()
			option = 'default'
			execute_default()
	elif len(sys.argv) > 2:
		initiate()
		option = sys.argv[2]
		if option == 'pre': execute_pre()
# 		elif option ==









# 	d.generate_alldata_dic()
# 	d.print_alldata_csv()
# 	d.generate_allposes_dic()
# 	print(d.alldata_dic)
# 	for x, y in d.allposes_dic.items():
# 		print(x)
# 		print(y.key)
# 	print(d.bs_key_list_dic)
# 	for bs, key_list in d.bs_key_list_dic.items():
# 		print("{}: {}".format(bs, len(key_list)))

# 	d.analyze_threshold()


# 	d.data_summary_by_lig_csv()
# 	d.make_bs_lig_E_list_dic()
# 	d.make_bs_lig_stdev_dic()
# 	print(d.bs_lig_stdev_dic)

# 	for x, y in d.alldata_dic.items():
# # 		print(x)
# # 		print(y)
# 		for yy in y: print(yy)
# 		cr()


# 	p = Pose(key = 'p1_ne_m2')
# 	p = Pose(dock = 'a1', lig = 'ghvbjnm', model = 6)
# 		for yy in y:
# 			print(yy)
# 	d.data_summary_by_lig_csv()

if __name__ == "__main__": main()
